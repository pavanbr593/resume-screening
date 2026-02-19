import json
import io
from typing import List, Dict, Any
from scoring_engine import recommendation_label


def generate_excel_report(candidates: List[Dict[str, Any]], job_title: str) -> bytes:
    """Generate an Excel ranking report for all candidates."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranking Report"

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Resume Screening Report — {job_title}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Rank", "Filename", "Final Score", "Skill Score",
        "Experience Score", "Semantic Score", "Quality Score",
        "Experience (yrs)", "Recommendation"
    ]
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    # Sort by final score
    sorted_candidates = sorted(
        [c for c in candidates if c.get("final_score") is not None],
        key=lambda x: x["final_score"],
        reverse=True
    )

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for rank, c in enumerate(sorted_candidates, 1):
        row = header_row + rank
        score = c.get("final_score", 0)
        rec = recommendation_label(score)

        values = [
            rank,
            c["filename"],
            score,
            c.get("skill_score", 0),
            c.get("experience_score", 0),
            c.get("semantic_score", 0),
            c.get("quality_score", 0),
            c.get("experience_years", 0),
            rec,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Color code by recommendation
        color = _rec_color(rec)
        for col in range(1, len(values) + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, header_row + len(sorted_candidates) + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # Skills sheet
    ws2 = wb.create_sheet("Skill Breakdown")
    ws2.append(["Filename", "Matched Skills", "Missing Skills"])
    for c in sorted_candidates:
        matched = json.loads(c.get("matched_skills") or "[]")
        missing = json.loads(c.get("missing_skills") or "[]")
        ws2.append([c["filename"], ", ".join(matched), ", ".join(missing)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _rec_color(rec: str) -> str:
    mapping = {
        "Strongly Recommended": "DCFCE7",
        "Recommended": "FEF9C3",
        "Consider": "FEF3C7",
        "Not Recommended": "FEE2E2",
    }
    return mapping.get(rec, "FFFFFF")


def generate_candidate_pdf_report(candidate: Dict[str, Any], job_title: str) -> bytes:
    """Generate a simple text-based PDF report for a single candidate."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        raise ImportError("reportlab is required: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.75 * inch)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"Candidate Report — {job_title}", styles["Title"]))
    story.append(Paragraph(f"File: {candidate['filename']}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # Score table
    score_data = [
        ["Metric", "Score"],
        ["Final Score", f"{candidate.get('final_score', 0):.1f}"],
        ["Skill Score", f"{candidate.get('skill_score', 0):.1f}"],
        ["Experience Score", f"{candidate.get('experience_score', 0):.1f}"],
        ["Semantic Score", f"{candidate.get('semantic_score', 0):.1f}"],
        ["Quality Score", f"{candidate.get('quality_score', 0):.1f}"],
        ["Experience (years)", f"{candidate.get('experience_years', 0):.1f}"],
        ["Recommendation", recommendation_label(candidate.get("final_score", 0))],
    ]
    t = Table(score_data, colWidths=[3 * inch, 2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.2 * inch))

    # Skills
    matched = json.loads(candidate.get("matched_skills") or "[]")
    missing = json.loads(candidate.get("missing_skills") or "[]")
    story.append(Paragraph("Matched Skills", styles["Heading2"]))
    story.append(Paragraph(", ".join(matched) or "None", styles["Normal"]))
    story.append(Spacer(1, 0.1 * inch))
    story.append(Paragraph("Missing Skills", styles["Heading2"]))
    story.append(Paragraph(", ".join(missing) or "None", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # AI Summary
    if candidate.get("ai_summary"):
        story.append(Paragraph("AI Summary", styles["Heading2"]))
        for line in candidate["ai_summary"].split("\n"):
            story.append(Paragraph(line, styles["Normal"]))
        story.append(Spacer(1, 0.1 * inch))

    doc.build(story)
    buf.seek(0)
    return buf.read()
